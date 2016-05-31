# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, colors, PatternFill
import math
# from app import create_app
# from app.models import Part


alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
border = Border(left=Side(border_style='thin', color='FF000000'),
	right=Side(border_style='thin', color='FF000000'),
	top=Side(border_style='thin', color='FF000000'),
	bottom=Side(border_style='thin', color='FF000000'))
font_big = Font(size=15)
font = Font(size=7.5)
# base_row = 1

bfill = PatternFill(start_color='CCFFCC',
                end_color='CCFFCC',
                fill_type='solid')


def make_cell(ws, cell, text):
	ws[cell] = text
	ws[cell].alignment = alignment 
	ws[cell].border = border
	ws[cell].font = font


def create_head(ws, base_row, headname):
	ws['A' + str(base_row)] = headname
	ws['A' + str(base_row)].alignment = alignment 
	ws['A' + str(base_row)].border = border
	ws['A' + str(base_row)].font = font_big
	ws.merge_cells('A' + str(base_row) +':L' + str(base_row+1))
	# base_row = base_row + 2


def create_columnn_head(ws, base_row):
	make_cell(ws, 'A'+str(base_row), '序号')
	make_cell(ws, 'B'+str(base_row), '图纸信息')
	ws.merge_cells('B' + str(base_row) +':C' + str(base_row))
	make_cell(ws, 'D'+str(base_row), '材料成本')
	ws.merge_cells('D' + str(base_row) +':E' + str(base_row))
	make_cell(ws, 'F'+str(base_row), '加工成本')
	ws.merge_cells('F' + str(base_row) +':J' + str(base_row))
	make_cell(ws, 'K'+str(base_row), '总计')
	make_cell(ws, 'L'+str(base_row), '备注')
	# base_row = base_row + 1


def create_part(ws, base_row, part):
	make_cell(ws, 'A'+str(base_row), part.id)
	ws.merge_cells('A' + str(base_row) +':A' + str(base_row+16))	
	make_cell(ws, 'B'+str(base_row), "工件名称")
	ws.merge_cells('B' + str(base_row) +':B' + str(base_row+3))
	make_cell(ws, 'B'+str(base_row+4), "图号")
	ws.merge_cells('B' + str(base_row+4) +':B' + str(base_row+6))
	make_cell(ws, 'B'+str(base_row+7), "精度要求")
	ws.merge_cells('B' + str(base_row+7) +':B' + str(base_row+9))
	make_cell(ws, 'B'+str(base_row+10), "工件类型")
	ws.merge_cells('B' + str(base_row+10) +':B' + str(base_row+12))
	make_cell(ws, 'B'+str(base_row+13), "材料类型")
	ws.merge_cells('B' + str(base_row+13) +':B' + str(base_row+15))
	make_cell(ws, 'B'+str(base_row+16), "")	
	make_cell(ws, 'C'+str(base_row), part.name)
	ws.merge_cells('C' + str(base_row) +':C' + str(base_row+3))
	make_cell(ws, 'C'+str(base_row+4), part.drawing_number)
	ws.merge_cells('C' + str(base_row+4) +':C' + str(base_row+6))
	make_cell(ws, 'C'+str(base_row+7), part.measure)
	ws['C'+str(base_row+7)].fill = bfill
	ws.merge_cells('C' + str(base_row+7) +':C' + str(base_row+9))
	make_cell(ws, 'C'+str(base_row+10), part.part_type)
	ws['C'+str(base_row+10)].fill = bfill
	ws.merge_cells('C' + str(base_row+10) +':C' + str(base_row+12))
	make_cell(ws, 'C'+str(base_row+13), part.material_type)
	ws['C'+str(base_row+13)].fill = bfill
	ws.merge_cells('C' + str(base_row+13) +':C' + str(base_row+15))
	make_cell(ws, 'C'+str(base_row+16), "")	
	make_cell(ws, 'D'+str(base_row), "材料")
	ws.merge_cells('D' + str(base_row) +':D' + str(base_row+3))
	make_cell(ws, 'D'+str(base_row+4), "数量")
	ws.merge_cells('D' + str(base_row+4) +':D' + str(base_row+6))
	make_cell(ws, 'D'+str(base_row+7), "下料尺寸")
	ws.merge_cells('D' + str(base_row+7) +':D' + str(base_row+9))
	make_cell(ws, 'D'+str(base_row+10), "理论重量")
	ws.merge_cells('D' + str(base_row+10) +':D' + str(base_row+12))
	make_cell(ws, 'D'+str(base_row+13), "材料单价")
	ws.merge_cells('D' + str(base_row+13) +':D' + str(base_row+14))
	make_cell(ws, 'D'+str(base_row+15), "材料金额")
	ws.merge_cells('D' + str(base_row+15) +':D' + str(base_row+16))
	make_cell(ws, 'E'+str(base_row), part.material)
	ws.merge_cells('E' + str(base_row) +':E' + str(base_row+3))
	ws['E'+str(base_row)].fill = bfill
	make_cell(ws, 'E'+str(base_row+4), part.number)
	ws.merge_cells('E' + str(base_row+4) +':E' + str(base_row+6))
	if part.material_type == "板材":
		make_cell(ws, 'E'+str(base_row+7), "L:{0}\nW:{1}\nH:{2}".format(part.size1, part.size2, part.size3))
		ws.merge_cells('E' + str(base_row+7) +':E' + str(base_row+9))
		make_cell(ws, 'E'+str(base_row+10), part.weight)
		ws.merge_cells('E' + str(base_row+10) +':E' + str(base_row+12))
	if part.material_type == "管材":
		make_cell(ws, 'E'+str(base_row+7), "D:{0}\nd:{1}\nL:{2}".format(part.size1, part.size2, part.size3))
		ws.merge_cells('E' + str(base_row+7) +':E' + str(base_row+9))
		make_cell(ws, 'E'+str(base_row+10), part.weight)
		ws.merge_cells('E' + str(base_row+10) +':E' + str(base_row+12))
	if part.material_type == "棒材":
		make_cell(ws, 'E'+str(base_row+7), "Φ:{0}\nL:{1}".format(part.size1, part.size2))
		ws.merge_cells('E' + str(base_row+7) +':E' + str(base_row+9))
		make_cell(ws, 'E'+str(base_row+10), part.weight)
		ws.merge_cells('E' + str(base_row+10) +':E' + str(base_row+12))
	make_cell(ws, 'E'+str(base_row+13), part.material_price)
	ws.merge_cells('E' + str(base_row+13) +':E' + str(base_row+14))
	make_cell(ws, 'E'+str(base_row+15), "=E{0}*E{1}".format(base_row+10, base_row+13))
	ws.merge_cells('E' + str(base_row+15) +':E' + str(base_row+16))
	make_cell(ws, 'F'+str(base_row), "加工类型")
	make_cell(ws, 'F'+str(base_row+1), "粗加工")
	ws.merge_cells('F' + str(base_row+1) +':F' + str(base_row+7))
	make_cell(ws, 'F'+str(base_row+8), "精加工")
	ws.merge_cells('F' + str(base_row+8) +':F' + str(base_row+15))
	make_cell(ws, 'F'+str(base_row+16), "零件加工费单价（含热处理费）")
	ws.merge_cells('F' + str(base_row+16) +':I' + str(base_row+16))
	make_cell(ws, 'G'+str(base_row), "加工流程")
	make_cell(ws, 'G'+str(base_row+1), part.rude_process1)
	make_cell(ws, 'G'+str(base_row+2), part.rude_process2)
	make_cell(ws, 'G'+str(base_row+3), part.rude_process3)
	make_cell(ws, 'G'+str(base_row+4), part.rude_process4)
	make_cell(ws, 'G'+str(base_row+5), part.rude_process5)
	make_cell(ws, 'G'+str(base_row+6), part.rude_process6)
	make_cell(ws, 'G'+str(base_row+7), part.rude_process7)
	make_cell(ws, 'G'+str(base_row+8), part.fine_process1)
	make_cell(ws, 'G'+str(base_row+9), part.fine_process2)
	make_cell(ws, 'G'+str(base_row+10), part.fine_process3)
	make_cell(ws, 'G'+str(base_row+11), part.fine_process4)
	make_cell(ws, 'G'+str(base_row+12), part.fine_process5)
	make_cell(ws, 'G'+str(base_row+13), part.fine_process6)
	make_cell(ws, 'G'+str(base_row+14), part.fine_process7)
	make_cell(ws, 'G'+str(base_row+15), part.fine_process8)
	make_cell(ws, 'H'+str(base_row), "耗时")
	make_cell(ws, 'H'+str(base_row+1), part.rude_process1_time if part.rude_process1_time else "")
	make_cell(ws, 'H'+str(base_row+2), part.rude_process2_time if part.rude_process2_time else "")
	make_cell(ws, 'H'+str(base_row+3), part.rude_process3_time if part.rude_process3_time else "")
	make_cell(ws, 'H'+str(base_row+4), part.rude_process4_time if part.rude_process4_time else "")
	make_cell(ws, 'H'+str(base_row+5), part.rude_process5_time if part.rude_process5_time else "")
	make_cell(ws, 'H'+str(base_row+6), part.rude_process6_time if part.rude_process6_time else "")
	make_cell(ws, 'H'+str(base_row+7), part.rude_process7_time if part.rude_process7_time else "")
	make_cell(ws, 'H'+str(base_row+8), part.fine_process1_time if part.fine_process1_time else "")
	make_cell(ws, 'H'+str(base_row+9), part.fine_process2_time if part.fine_process2_time else "")
	make_cell(ws, 'H'+str(base_row+10), part.fine_process3_time if part.fine_process3_time else "")
	make_cell(ws, 'H'+str(base_row+11), part.fine_process4_time if part.fine_process4_time else "")
	make_cell(ws, 'H'+str(base_row+12), part.fine_process5_time if part.fine_process5_time else "")
	make_cell(ws, 'H'+str(base_row+13), part.fine_process6_time if part.fine_process6_time else "")
	make_cell(ws, 'H'+str(base_row+14), part.fine_process7_time if part.fine_process7_time else "")
	make_cell(ws, 'H'+str(base_row+15), part.fine_process8_time if part.fine_process8_time else "")
	make_cell(ws, 'I'+str(base_row), "单位成本")
	make_cell(ws, 'I'+str(base_row+1), part.rude_process1_price if part.rude_process1_price else "")
	make_cell(ws, 'I'+str(base_row+2), part.rude_process2_price if part.rude_process2_price else "")
	make_cell(ws, 'I'+str(base_row+3), part.rude_process3_price if part.rude_process3_price else "")
	make_cell(ws, 'I'+str(base_row+4), part.rude_process4_price if part.rude_process4_price else "")
	make_cell(ws, 'I'+str(base_row+5), part.rude_process5_price if part.rude_process5_price else "")
	make_cell(ws, 'I'+str(base_row+6), part.rude_process6_price if part.rude_process6_price else "")
	make_cell(ws, 'I'+str(base_row+7), part.rude_process7_price if part.rude_process7_price else "")
	make_cell(ws, 'I'+str(base_row+8), part.fine_process1_price if part.fine_process1_price else "")
	make_cell(ws, 'I'+str(base_row+9), part.fine_process2_price if part.fine_process2_price else "")
	make_cell(ws, 'I'+str(base_row+10), part.fine_process3_price if part.fine_process3_price else "")
	make_cell(ws, 'I'+str(base_row+11), part.fine_process4_price if part.fine_process4_price else "")
	make_cell(ws, 'I'+str(base_row+12), part.fine_process5_price if part.fine_process5_price else "")
	make_cell(ws, 'I'+str(base_row+13), part.fine_process6_price if part.fine_process6_price else "")
	make_cell(ws, 'I'+str(base_row+14), part.fine_process7_price if part.fine_process7_price else "")
	make_cell(ws, 'I'+str(base_row+15), part.fine_process8_price if part.fine_process8_price else "")
	make_cell(ws, 'J'+str(base_row), "单项总计")
	make_cell(ws, 'J'+str(base_row+1), "=H{0}*I{0}".format(base_row+1) if part.rude_process1 else "")
	make_cell(ws, 'J'+str(base_row+2), "=H{0}*I{0}".format(base_row+2) if part.rude_process2 else "")
	make_cell(ws, 'J'+str(base_row+3), "=H{0}*I{0}".format(base_row+3) if part.rude_process3 else "")
	make_cell(ws, 'J'+str(base_row+4), "=H{0}*I{0}".format(base_row+4) if part.rude_process4 else "")
	make_cell(ws, 'J'+str(base_row+5), "=H{0}*I{0}".format(base_row+5) if part.rude_process5 else "")
	make_cell(ws, 'J'+str(base_row+6), "=H{0}*I{0}".format(base_row+6) if part.rude_process6 else "")
	make_cell(ws, 'J'+str(base_row+7), "=H{0}*I{0}".format(base_row+7) if part.rude_process7 else "")
	make_cell(ws, 'J'+str(base_row+8), "=H{0}*I{0}".format(base_row+8) if part.fine_process1 else "")
	make_cell(ws, 'J'+str(base_row+9), "=H{0}*I{0}".format(base_row+9) if part.fine_process2 else "")
	make_cell(ws, 'J'+str(base_row+10), "=H{0}*I{0}".format(base_row+10) if part.fine_process3 else "")
	make_cell(ws, 'J'+str(base_row+11), "=H{0}*I{0}".format(base_row+11) if part.fine_process4 else "")
	make_cell(ws, 'J'+str(base_row+12), "=H{0}*I{0}".format(base_row+12) if part.fine_process5 else "")
	make_cell(ws, 'J'+str(base_row+13), "=H{0}*I{0}".format(base_row+13) if part.fine_process6 else "")
	make_cell(ws, 'J'+str(base_row+14), "=H{0}*I{0}".format(base_row+14) if part.fine_process7 else "")
	make_cell(ws, 'J'+str(base_row+15), "=H{0}*I{0}".format(base_row+15) if part.fine_process8 else "")
	make_cell(ws, 'J'+str(base_row+16), "=SUM(J{0}:J{1})".format(base_row+1, base_row+15))
	make_cell(ws, 'K'+str(base_row), "=(J{0}+E{1})*E{2}".format(base_row+16, base_row+15, base_row+4))
	ws.merge_cells('K' + str(base_row) +':K' + str(base_row+16))
	ws['K' + str(base_row)].number_format = '"¥"0.00'
	make_cell(ws, 'L'+str(base_row), part.commit)
	ws.merge_cells('L' + str(base_row) +':L' + str(base_row+16))		



import os
from flask import Flask
from flask.ext.sqlalchemy import SQLAlchemy
basedir = os.path.abspath(os.path.dirname(__file__))
app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(basedir, 'data_dev.sqlite')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = True
db = SQLAlchemy(app)

class Part(db.Model):
    __tablename__ = 'Parts'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(64), unique=True)
    drawing_number = db.Column(db.String(128), unique=True)
    measure = db.Column(db.String(64))
    part_type = db.Column(db.String(64))
    material_type = db.Column(db.String(64))
    material = db.Column(db.String(64))
    number = db.Column(db.Integer)
    size1 = db.Column(db.Float)
    size2 = db.Column(db.Float)
    size3 = db.Column(db.Float)
    weight = db.Column(db.Float)
    commit = db.Column(db.String(256))
    material_price = db.Column(db.Float)
    material_total_price = db.Column(db.Float)
    rude_process1 = db.Column(db.String(64))
    rude_process2 = db.Column(db.String(64))
    rude_process3 = db.Column(db.String(64))
    rude_process4 = db.Column(db.String(64))
    rude_process5 = db.Column(db.String(64))
    rude_process6 = db.Column(db.String(64))
    rude_process7 = db.Column(db.String(64))
    fine_process1 = db.Column(db.String(64))
    fine_process2 = db.Column(db.String(64))
    fine_process3 = db.Column(db.String(64))
    fine_process4 = db.Column(db.String(64))
    fine_process5 = db.Column(db.String(64))
    fine_process6 = db.Column(db.String(64))
    fine_process7 = db.Column(db.String(64))
    fine_process8 = db.Column(db.String(64))

    rude_process1_time = db.Column(db.Float)
    rude_process2_time = db.Column(db.Float)
    rude_process3_time = db.Column(db.Float)
    rude_process4_time = db.Column(db.Float)
    rude_process5_time = db.Column(db.Float)
    rude_process6_time = db.Column(db.Float)
    rude_process7_time = db.Column(db.Float)
    fine_process1_time = db.Column(db.Float)
    fine_process2_time = db.Column(db.Float)
    fine_process3_time = db.Column(db.Float)
    fine_process4_time = db.Column(db.Float)
    fine_process5_time = db.Column(db.Float)
    fine_process6_time = db.Column(db.Float)
    fine_process7_time = db.Column(db.Float)
    fine_process8_time = db.Column(db.Float)

    rude_process1_price = db.Column(db.Float)
    rude_process2_price = db.Column(db.Float)
    rude_process3_price = db.Column(db.Float)
    rude_process4_price = db.Column(db.Float)
    rude_process5_price = db.Column(db.Float)
    rude_process6_price = db.Column(db.Float)
    rude_process7_price = db.Column(db.Float)
    fine_process1_price = db.Column(db.Float)
    fine_process2_price = db.Column(db.Float)
    fine_process3_price = db.Column(db.Float)
    fine_process4_price = db.Column(db.Float)
    fine_process5_price = db.Column(db.Float)
    fine_process6_price = db.Column(db.Float)
    fine_process7_price = db.Column(db.Float)
    fine_process8_price = db.Column(db.Float)

    rude_process1_total_price = db.Column(db.Float)
    rude_process2_total_price = db.Column(db.Float)
    rude_process3_total_price = db.Column(db.Float)
    rude_process4_total_price = db.Column(db.Float)
    rude_process5_total_price = db.Column(db.Float)
    rude_process6_total_price = db.Column(db.Float)
    rude_process7_total_price = db.Column(db.Float)
    fine_process1_total_price = db.Column(db.Float)
    fine_process2_total_price = db.Column(db.Float)
    fine_process3_total_price = db.Column(db.Float)
    fine_process4_total_price = db.Column(db.Float)
    fine_process5_total_price = db.Column(db.Float)
    fine_process6_total_price = db.Column(db.Float)
    fine_process7_total_price = db.Column(db.Float)
    fine_process8_total_price = db.Column(db.Float)

    single_process_price = db.Column(db.Float)
    total_price = db.Column(db.Float)

    def __repr__(self):
        return "Id: {0} name {1} drawing_number {2} number {3}".format(self.id, self.name, self.drawing_number, self.number)



if __name__ == '__main__':
	wb = Workbook()
	ws = wb.active
	for i in range(1, 10):
		row = ws.row_dimensions[i]
		row.height = 13.3
	ws.column_dimensions['A'].width = 4.11
	ws.column_dimensions['B'].width = 6.47
	ws.column_dimensions['C'].width = 6.47
	ws.column_dimensions['D'].width = 6.47
	ws.column_dimensions['E'].width = 6.47
	ws.column_dimensions['F'].width = 6.47
	ws.column_dimensions['G'].width = 6.47
	ws.column_dimensions['H'].width = 6.47
	ws.column_dimensions['I'].width = 6.47
	ws.column_dimensions['J'].width = 6.47
	ws.column_dimensions['K'].width = 8.79
	ws.column_dimensions['L'].width = 6.75
	base_row = 1
	create_head(ws, base_row, 'MAHLE 成本核价')
	base_row = base_row + 2
	create_columnn_head(ws, base_row)
	base_row = base_row + 1
	parts = Part.query.all()
	for part in parts:
		# print(part)
		create_part(ws, base_row, part)
		base_row = base_row + 17

	for i in range(1, len(parts)*17+3+1):
		row = ws.row_dimensions[i]
		row.height = 13.3
	base_row = 1
	wb.save('test.xlsx')
